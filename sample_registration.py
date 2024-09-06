from tkinter import *
import tkinter
from tkinter import messagebox 
from openpyxl.workbook import Workbook 
from openpyxl import load_workbook 

window=Tk()
window.geometry("500x900")
wb = Workbook()
wb = load_workbook('register.xlsx')
ws = wb['Sheet1']


label= Label(window,text="Register",font=("bold",20), fg='red', pady=(20))
label.pack()

label_name=Label(window,text="Name",font=("bold",14),pady=(20))
label_name.pack()

name=Entry(window, width=25, font=('Helvetica',14))
name.pack()

  
label_phone=Label(window,text="Phone",font=("bold",14),pady=(20))
label_phone.pack()
phone=Entry(window,width=25, font=('Helvetica',14))
phone.pack()

label_add=Label(window,text="Address",font=("bold",14),pady=(20))
label_add.pack()
add1=Entry(window,width=25, font=('Helvetica',14))
add1.pack()

addbtn=Button(window,text="Add",font=("Helevetica",11,"bold"),command=lambda:save_data())
   
def save_data():
    for each_cell in range(2,(ws.max_row)+1):
        if ((name.get() == ws['A' + str(each_cell)].value) or (phone.get() == ws['B' + str(each_cell)].value)):
            Found = True;
            break;
        else:
            Found = False
    if(Found == True):
        messagebox.showinfo("ERROR","DATA ALREADY EXISTED")
    else:
        messagebox.showinfo("SUCCESS","DATA SAVED SUCCESSFULLY")
        lastRow = str((ws.max_row) + 1)
        ws['A' + lastRow] = name.get() 
        ws['B' + lastRow] = phone.get()
        ws['C' + lastRow] = add1.get()

        clear_entries()
    wb.save('register.xlsx')

def search_data():
    for each_cell in range(2, (ws.max_row) + 1):
        if ((name.get() == ws['A' + str(each_cell)].value) or (phone.get() == ws['B' + str(each_cell)].value)):
            Found = True;
            cell_address = str(each_cell)
            break
        else:
            Found = False
    if(Found == True):
        messagebox.info("FOUND","DATA EXIST IN " + cell_address)
        phone.insert(0,ws['B' + cell_address].value)
        add1.insert(0,ws['C' + cell_address].value)

def delete_data():
    for each_cell in range(2, (ws.max_row) + 1):
        if (name.get() == ws['A' + str(each_cell)].value) or (phone.get() == ws['B' + str(each_cell)].value):
            Found = True;
            cell_address = each_cell
            break;
        else:
            Found = False
    if(Found == True):
        ws.delete_rows(cell_address)
        messagebox.info("INFORMATION DELETED")
        clear_entries()
    wb.save('register.xlsx')

def clear_entries():
    name.delete(0, END)
    phone.delete(0, END)
    add1.delete(0, END)


dsearch=Button(window,text="Search",font=("Helevetica",11,"bold"), command=lambda:search_data())
show_Result = Button(window,text="SHOW RESULT",font=("Helevetica",11,"bold"), command=lambda:show_result)
delete = Button(window,text="Delete",font=("Helevetica",11,"bold"), command=lambda:delete_data)

column_A = ws['A']
column_B = ws['B']
column_C = ws['C']

results_show = Label(window,font=('Helvetica',14))

def show_result():
    results_list = ""
    for each_cell in column_A:
        print(each_cell.value)
        results_list = results_list + each_cell.value    
        results_list = f'{results_list + str(each_cell.value)}\n'
    results_show.configure(text=results_list)

addbtn.pack()
dsearch.pack()
show_Result.pack()
delete.pack()

res=Label(window,text="Result",font=("bold",17,"underline"),pady=20)
res.pack()

window.mainloop()

