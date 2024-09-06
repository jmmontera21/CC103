from http.client import CannotSendHeader
import openpyxl
import tkinter  
from tkinter import messagebox
from tkinter import simpledialog
from tkinter import *


wb=openpyxl.load_workbook('register1.xlsx')
ws = wb['Sheet1']

window=Tk()

window.geometry("500x900")

label=tkinter.Label(window,text="Register",font=("bold",20), fg='blue', pady=(20))
label.pack()


label_name=tkinter.Label(window,text="Name",font=("bold",14),pady=(20))
label_name.pack()

name=tkinter.Entry(window,width=20,font=('Helvetica',14))
name.pack()


label_phone=tkinter.Label(window,text="Phone",font=("bold",14),pady=(20))
label_phone.pack()
phone=tkinter.Entry(window,width=20,font=('Helvetica',14))
phone.pack()


label_add=tkinter.Label(window,text="Address",font=("bold",14),pady=(20))
label_add.pack()
add1=tkinter.Entry(window,width=20,font=('Helvetica',14))
add1.pack()

def search():
    for each_cell in range(2,(ws.max_row)+1):
        if (name.get() == ws['A'+str(each_cell)].value) or (phone.get() == ws['B'+str(each_cell)].value):  
            Found =  True;
            cell_address = str(each_cell)
            break;
        else:
            Found=False
    if(Found == True):
        messagebox.showinfo("Found","Data Exist IN" + cell_address)
        phone.insert(0,ws['B'+cell_address].value)
        add1.insert(0,ws['C'+cell_address].value) 

def show_result():
    result_name.config(text="")
    result_phone.config(text="")
    result_addr.config(text="")
    result_add.config(text="")
    

    for each_cell in range(2,(ws.max_row)+1):
         if((name.get()==ws['A'+str(each_cell)].value) or (phone.get()==ws['B'+str(each_cell)].value)):
             Found=True
             break

         else:
            Found = False
            
            
    if (Found==True):
        result_name.config(text="Name: "+str(ws['A'+str(each_cell)].value))
        result_phone.config(text="Phone: "+str(ws['B'+str(each_cell)].value))
        result_addr.config(text="Address: "+str(ws['C'+str(each_cell)].value))
    
    else:
        result_add.config(text="No record found.")       
        

def add():
    for each_cell in range(2,(ws.max_row)+1):
        if((name.get()==ws['A'+str(each_cell)].value) and (phone.get()==ws['B'+str(each_cell)].value)):
          
            Found=True
            break

        else:
            Found = False
            print(ws['A'+str(each_cell)])

    if(Found==True):
         messagebox.showinfo("Error", "Name Exists!")
                
    else:
        messagebox.showinfo("Success", "Record Added!")
        lastRow=str((ws.max_row)+1)
        ws['A'+lastRow]=name.get()
        ws['B'+lastRow]=phone.get()
        ws['C'+lastRow]=add1.get()
        clear_entry()
    wb.save('register1.xlsx')

def clear_entry():
    name.delete(0,END)
    phone.delete(0,END)
    add1.delete(0,END)

asearch=tkinter.Button(window,text="Add",command=add, bg='yellow', font=("Helevetica",11,"bold"))



def edit():
    for each_cell in range(2,(ws.max_row)+1):
        if((name.get()==ws['A'+str(each_cell)].value) or (phone.get()==ws['B'+str(each_cell)].value)):
            Found=True
            break

        else:
            Found = False

    if(Found==True):
             
            top=Toplevel()
            top.geometry("500x600")
            top.title('Change of Information')
            A=""
          
            
            def nam():
                global A
                if(var1.get()==1):
                    na=name.get()
                    s.set(na)
                    A=na
                elif(var1.get()==0):
                    s.set("")
            
                
            def ph():
                if(var2.get()==1):
                    ph=phone.get()
                    t.set(ph)
                    B=ph
                elif(var2.get()==0):
                    t.set("")

            def ad():
                if(var3.get()==1):
                    add=add1.get()
                    u.set(add)
                    C=add
                elif(var3.get()==0):
                    u.set("")

                
            label_name=tkinter.Label(top,text="New Name",font=("bold",14),pady=(20))
            label_name.pack()

            s = StringVar()
            t = StringVar()
            u = StringVar()


            
            name1=Entry(top, width=25, font=('Helvetica',14), textvariable=s)
            name1.pack()

            var1 = IntVar()
            c1=Checkbutton(top, text="same as before",variable=var1,command=nam)
            c1.pack()
            
            label_phone=tkinter.Label(top,text="New phone",font=("bold",14),pady=(20))
            label_phone.pack()

            phone1=Entry(top, width=25, font=('Helvetica',14), textvariable=t)
            phone1.pack()

            var2 = IntVar()
            c2=Checkbutton(top, text="same as before", variable=var2,command=ph)
            c2.pack()
                
            label_add=tkinter.Label(top,text="New Address",font=("bold",14),pady=(20))
            label_add.pack()

            add2=Entry(top, width=25, font=('Helvetica',14), textvariable=u)
            add2.pack()

            var3 = IntVar()
            c3=Checkbutton(top, text="same as before", variable=var3,command=ad)
            c3.pack()
            
            def update():
                print(ws['A'+str(each_cell)])
                ws['A'+str(each_cell)] = name1.get()
                ws['B'+str(each_cell)] = phone1.get()
                ws['C'+str(each_cell)] = add2.get()
                messagebox.showinfo("The information was updated successfully")
                clear_entries()
                wb.save('register1.xlsx')


            Bupdate=tkinter.Button(top,text="Update",command=update, bg='red', font=("Helevetica",11,"bold"))
            Bupdate.pack()    
            
            top.mainloop()
        
        
                
    else:
        result_add.config(text="no record found to update!")


def delete():
    for i in range(2,(ws.max_row)+1):
        if (name.get() == ws['A'+str(i)].value) or (phone.get() == ws['B'+str(i)].value):  
            Found =  True;
            cell_address = (i)
            break;
        else:
            Found=False
    if(Found == True):
        ws.delete_rows(cell_address)
        messagebox.showinfo("Info","Data Deleted")
        clear_entries()
    wb.save('register1.xlsx')  

def clear_entries():
    name.delete(0, END)
    phone.delete(0, END)
    add1.delete(0, END)

var1 = IntVar()
Checkbutton(window, text="no change", variable=var1)
edit=tkinter.Button(window,text="Edit",command=edit, bg='pink', font=("Helevetica",11,"bold"))
csearch =tkinter.Button(window,text="Search",command=search, bg='lightblue', font=("Helevetica",11,"bold"))
dsearch=tkinter.Button(window,text="Delete",command=delete, bg='orange', font=("Helevetica",11,"bold"))
bsearch=tkinter.Button(window,text="Show Result",command=show_result, bg='green', font=("Helevetica",11,"bold"))



asearch.pack()
edit.pack()
dsearch.pack()
csearch.pack()
bsearch.pack()

res=tkinter.Label(window,text="Result",font=("bold",17,"underline"),pady=20)
res.pack()



result_name=tkinter.Label(window,font=("bold",15))
result_name.pack()

result_phone=tkinter.Label(window,font=("bold",15))
result_phone.pack()

result_addr=tkinter.Label(window,font=("bold",15))
result_addr.pack()

result_add=tkinter.Label(window,font=("bold",15))
result_add.pack()


window.mainloop()
