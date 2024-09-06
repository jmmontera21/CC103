import tkinter 
import openpyxl
from tkinter import *
import os
from tkinter import messagebox
from tkinter import simpledialog
from openpyxl.workbook import Workbook
from openpyxl import load_workbook

#NOTE: ONLY IF NEEDED 
#os.chdir('C:\\Users\\Timothy Carl\\Visual Studio Code')
#os.getcwd()  

window=Tk()
window.geometry("500x650")

wb = Workbook()
wb = load_workbook('register.xlsx')
ws = wb.active
#ws = wb['Sheet1']

label=tkinter.Label(window,text="Register",font=("bold",20))
label.pack()

label_name=tkinter.Label(window,text="Name",font=("bold",14),pady=(20))
label_name.pack()
name=tkinter.Entry(window, width=10, font=('Helvetica',14))
name.pack()

  
label_phone=Label(window,text="Phone",font=("bold",14),pady=(20))
label_phone.pack()
phone=Entry(window,width=10, font=('Helvetica',14))
phone.pack()

label_add=Label(window,text="Address",font=("bold",14),pady=(20))
label_add.pack()
add1=Entry(window,width=10, font=('Helvetica',14))
add1.pack()

var1 = IntVar()
Checkbutton(window, text="no change", variable=var1)

searchBtn=Button(window, text="Search",font=("Helvetica",11,"bold"), command=lambda:search_data())
def search_data():
    name.config(text="")
    phone.config(text="")
    add1.config(text="")
    result_add.config(text="")


    for each_cell in range(2, (ws.max_row) + 1):
        if (name.get() == ws ['A'+str(each_cell)].value) or (phone.get() == ws ['B'+str(each_cell)].value) or (add1.get() == ws ['C'+str(each_cell)].value):
            Found = True;
            cell_address = str(each_cell)
            break;
        else:
            Found=False
    if(Found == True):
        messagebox.showinfo("INFO","DATA ALREADY EXISTS IN ROW " + cell_address )
        phone.insert(0, ws['B'+cell_address].value)
        add1.insert(0, ws['C'+cell_address].value)
    else:
        result_add.config(text="No record found.") 


dsearch=Button(window,text="Delete",font=("Helevetica",11,"bold"),command=lambda:delete_data())
def delete_data():
    for each_cell in range(2, (ws.max_row) + 1):
        if (name.get() == ws ['A'+str(each_cell)].value) or (phone.get() == ws ['B'+str(each_cell)].value):
            Found = True;
            cell_address = each_cell
            break;
        else:
            Found=False
    if(Found == True):
        ws.delete_rows(cell_address)
        messagebox.showinfo("INFO", "DATA DELETED")
        clear_entries()
    wb.save('register.xlsx')

def clear_entries():
    name.delete(0,END)
    phone.delete(0,END)
    add1.delete(0,END)


edit=Button(window,text="Edit",font=("Helevetica",11,"bold"),command=lambda:edit_data())
def edit_data():
    for each_cell in range(2, (ws.max_row) + 1):
        if ((name.get() == ws ['A'+str(each_cell)].value) or (phone.get() == ws ['B'+str(each_cell)].value) or (add1.get() == ws ['C'+str(each_cell)].value)):
            Found = True
            
            break
        else:
            Found=False
    if(Found == True):
        edit_form = Toplevel()
        edit_form.geometry('400x500')


        name_excel = StringVar()
        phone_excel = StringVar()
        addr_excel = StringVar()

        edit_form.title('EDIT DATA FROM EXCEL')
        edit_label = Label(edit_form, text="Edit Form ", font=('Helvetica',20))
        edit_label.pack()
        
        #NAME
        name_lbl=Label(edit_form,text="New Name",font=("bold",14),pady=(20))
        name_lbl.pack()
        name_txt=Entry(edit_form, textvariable=name_excel, width=10, font=('Helvetica',14))
        name_txt.pack()
        name_choice = IntVar()
        name_chk = Checkbutton(edit_form, text="Same as before", variable=name_choice, onvalue=1, offvalue=0, command=lambda:name_exist())
        name_chk.pack()

        #PHONE
        phone_lbl=Label(edit_form,text="New Phone",font=("bold",14),pady=(20))
        phone_lbl.pack()
        phone_txt=Entry(edit_form, textvariable=phone_excel, width=10, font=('Helvetica',14))
        phone_txt.pack()
        phone_choice = IntVar()
        phone_chk = Checkbutton(edit_form, text="Same as before", variable=phone_choice, onvalue=1, offvalue=0,command=lambda:phone_exist())
        phone_chk.pack()

        #ADDRESS
        add1_lbl=Label(edit_form,text="New Address",font=("bold",14),pady=(20))
        add1_lbl.pack()
        add1_txt=Entry(edit_form, textvariable=addr_excel, width=10, font=('Helvetica',14))
        add1_txt.pack()
        add1_choice = IntVar()
        add1_chk = Checkbutton(edit_form, text="Same as before", variable=add1_choice, onvalue=1, offvalue=0,command=lambda:addr_exist())
        add1_chk.pack()

        def name_exist():
            if name_choice.get()==1:
                nameOld=name.get()
                name_excel.set(nameOld)
            elif name_choice.get()==0:
                name_excel.set("")

        def phone_exist():
            if phone_choice.get()==1:
                phoneOld=phone.get()
                phone_excel.set(phoneOld)
            elif phone_choice.get()==0:
                phone_excel.set("")

        def addr_exist():
            if add1_choice.get()==1:
                addrOld=add1.get()
                addr_excel.set(addrOld)
            elif add1_choice.get()==0:
                addr_excel.set("")
        
        edit_btn = tkinter.Button(edit_form,text="UPDATE VALUE", width=20, height=2,font=("Arial",20),command=lambda:update)
        edit_btn.pack(padx=10, pady=9)
        
        def update():
            ws['A'+str(each_cell)].value = name_txt.get()
            ws['B'+str(each_cell)].value = phone_txt.get()
            ws['C'+str(each_cell)].value = add1_txt.get()
            
            messagebox.showinfo("SUCESS","DATA UPDATED SUCCESSFULLY")
            wb.save('register.xlsx')
    
    else:
        result_add.config(text="No record found to update!")
        clear_entries()
    edit_form.mainloop()


showResult = Button(window, text="SHOW RESULT", font=("Helevetica",11,"bold"), command=lambda:show_result())
def show_result():
    result_name.config(text="")
    result_phone.config(text="")
    result_addr.config(text="")
    result_add.config(text="")
    

    for each_cell in range(2,(ws.max_row)+1):
         if((name.get()==ws['A'+str(each_cell)].value) or (phone.get()==ws['B'+str(each_cell)].value)):
             Found=True
             break

         else:
            Found = False
            
            
    if (Found==True):
        result_name.config(text="Name: "+str(ws['A'+str(each_cell)].value))
        result_phone.config(text="Phone: "+str(ws['B'+str(each_cell)].value))
        result_addr.config(text="Address: "+str(ws['C'+str(each_cell)].value))
    
    else:
        result_add.config(text="No record found.")


addbtn=Button(window,text="Add",font=("Helevetica",11,"bold"),command=lambda:save_data())
def save_data():
    for each_cell in range(2, (ws.max_row) + 1):
        if (name.get() == ws ['A'+str(each_cell)].value) or (phone.get() == ws ['B'+str(each_cell)].value):
            Found = True;    
            break;
        else:
            Found=False
    if(Found == True):
        messagebox.showinfo("ERROR","DATA ALREADY EXISTS")
    else:
        lastRow = str(ws.max_row+1)
        ws['A' +lastRow] = name.get()
        ws['B' +lastRow] = phone.get()
        ws['C' +lastRow] = add1.get()

        clear_entries()
        messagebox.showinfo("SUCESS","DATA SAVED SUCCESSFULLY")
    wb.save('register.xlsx')

column_A = ws['A']
column_B = ws['B']
column_C = ws['C']

clearbtn=Button(window,text="Clear",font=("Helevetica",11,"bold"),command=lambda:clear_entries())

#auto design
addbtn.pack()
clearbtn.pack()
edit.pack()
dsearch.pack()
searchBtn.pack()
showResult.pack()

res=tkinter.Label(window,text="Result",font=("bold",17,"underline"),pady=20)
res.pack()

result_name=tkinter.Label(window,font=("bold",15))
result_name.pack()

result_phone=tkinter.Label(window,font=("bold",15))
result_phone.pack()

result_addr=tkinter.Label(window,font=("bold",15))
result_addr.pack()

result_add=tkinter.Label(window,font=("bold",15))
result_add.pack()

window.mainloop()