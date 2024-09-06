from audioop import add
import tkinter 
from tkinter import *
from tkinter import messagebox
from openpyxl.workbook import Workbook
from openpyxl import load_workbook

root = Tk()
root.title("Biodata")
root.geometry("1100x750")

wb = Workbook()
wb=load_workbook('biodata.xlsx')
ws = wb.active

label=Label(root,text="BIO DATA",font=("bold",15), fg='blue')
label.grid(row=0,column=1)

label_name=Label(root,text="Name :",font=("bold",14))
label_name.grid(row=1,column=0,padx=3,pady=3)
name=Entry(root,width=30,font=('Helvetica',14))
name.grid(row=1,column=1,padx=3,pady=3)

label_mobile=Label(root,text="Mobile :",font=("bold",14))
label_mobile.grid(row=2,column=0,padx=3,pady=3)
mobile=Entry(root,width=30,font=('Helvetica',14))
mobile.grid(row=2,column=1,padx=3,pady=3)

label_email=Label(root,text="Email id :",font=("bold",14))
label_email.grid(row=3,column=0,padx=3,pady=3)
email=Entry(root,width=30,font=('Helvetica',14))
email.grid(row=3,column=1,padx=3,pady=3)

label_father=Label(root,text="Father's Name :",font=("bold",14))
label_father.grid(row=4,column=0,padx=3,pady=3)
fathers_name=Entry(root,width=30,font=('Helvetica',14))
fathers_name.grid(row=4,column=1,padx=3,pady=3)

label_mother=Label(root,text="Mother's Name :",font=("bold",14))
label_mother.grid(row=5,column=0,padx=3,pady=3)
mothers_name=Entry(root,width=30,font=('Helvetica',14))
mothers_name.grid(row=5,column=1,padx=3,pady=3)

label_gender=Label(root,text="Gender :",font=("bold",14))
label_gender.grid(row=6,column=0,padx=3,pady=3)
gender=Entry(root,width=30,font=('Helvetica',14))
gender.grid(row=6,column=1,padx=3,pady=3)

label_birthdate=Label(root,text="Date of Birth :",font=("bold",14))
label_birthdate.grid(row=7,column=0,padx=3,pady=3)
birthdate=Entry(root,width=30,font=('Helvetica',14))
birthdate.grid(row=7,column=1,padx=3,pady=3)

label_maritalstatus=Label(root,text="Marital Status :",font=("bold",14))
label_maritalstatus.grid(row=8,column=0,padx=3,pady=3)
marital_status=Entry(root,width=30,font=('Helvetica',14))
marital_status.grid(row=8,column=1,padx=3,pady=3)

label_religion=Label(root,text="Religion :",font=("bold",14))
label_religion.grid(row=9,column=0,padx=3,pady=3)
religion=Entry(root,width=30,font=('Helvetica',14))
religion.grid(row=9,column=1,padx=3,pady=3)

label_language=Label(root,text="Languages Known :",font=("bold",14))
label_language.grid(row=10,column=0,padx=3,pady=3)
language=Entry(root,width=30,font=('Helvetica',14))
language.grid(row=10,column=1,padx=3,pady=3)

label_qualification=Label(root,text="Qualification :",font=("bold",14))
label_qualification.grid(row=11,column=0,padx=3,pady=3)
qualification=Entry(root,width=30,font=('Helvetica',14))
qualification.grid(row=11,column=1,padx=3,pady=3,ipady=20)

label_experience=Label(root,text="Experience :",font=("bold",14))
label_experience.grid(row=12,column=0,padx=3,pady=3)
experience=Entry(root,width=30,font=('Helvetica',14))
experience.grid(row=12,column=1,padx=3,pady=3,ipady=20)

label_address=Label(root,text="Address :",font=("bold",14))
label_address.grid(row=13,column=0,padx=3,pady=3)
address=Entry(root,width=30,font=('Helvetica',14))
address.grid(row=13,column=1,padx=3,pady=3)

label_place=Label(root,text="Place :",font=("bold",14))
label_place.grid(row=14,column=0,padx=3,pady=3)
place=Entry(root,width=20,font=('Helvetica',14))
place.grid(row=14,column=1,padx=3,pady=3,sticky=W)

label_date=Label(root,text="Date :",font=("bold",14))
label_date.grid(row=15,column=0,padx=3,pady=3)
date=Entry(root,width=20,font=('Helvetica',14))
date.grid(row=15,column=1,padx=3,pady=3,sticky=W)

button1=Button(root,text="Save",width=10,height=2,bg='yellow', font=("Helevetica",11,"bold"),command=lambda:save_data())
button1.place(x=20,y=620)

def save_data():
    for each_cell in range(2,(ws.max_row)+1):
        if (name.get() == ws['A'+str(each_cell)].value) or (mobile.get() == ws['B'+str(each_cell)].value):  
            Found =  True;
            break;
        else:
            Found = False
    if(Found ==True):
        messagebox.showinfo("ERROR","Data Already Exist")
    else:
        lastrow = str((ws.max_row)+1)
        ws['A'+lastrow] = name.get()
        ws['B'+lastrow] = mobile.get()
        ws['C'+lastrow] = email.get()   
        ws['D'+lastrow] = fathers_name.get()
        ws['E'+lastrow] = mothers_name.get()
        ws['F'+lastrow] = gender.get()   
        ws['G'+lastrow] = birthdate.get()
        ws['H'+lastrow] = marital_status.get()
        ws['I'+lastrow] = religion.get()   
        ws['J'+lastrow] = language.get()
        ws['K'+lastrow] = qualification.get()
        ws['L'+lastrow] = experience.get()   
        ws['M'+lastrow] = address.get()
        ws['N'+lastrow] = place.get()   
        ws['O'+lastrow] = date.get()


        messagebox.showinfo("SUCCESS","Data Saved Successfully")
        clear_entries()
    wb.save('biodata.xlsx')

button2=Button(root,text="Edit",width=10,height=2,bg='pink', font=("Helevetica",11,"bold"),command=lambda:edit())
button2.place(x=130,y=620)

def edit():
    for each_cell in range(2,(ws.max_row)+1):
        if((name.get()==ws['A'+str(each_cell)].value) or (mobile.get()==ws['B'+str(each_cell)].value)):
            Found=True
            break

        else:
            Found = False

    if(Found==True):
             
            top=Toplevel()
            top.geometry("900x600")
            top.title('Change of Information')
            A=""
          
            
            def nam():
                global A
                if(var1.get()==1):
                    na=name.get()
                    l.set(na)
                    A=na
                elif(var1.get()==0):
                    l.set("")
                
            def ph():
                if(var2.get()==1):
                    ph=mobile.get()
                    m.set(ph)
                    B=ph
                elif(var2.get()==0):
                    m.set("")

            def ad():
                if(var3.get()==1):
                    add=email.get()
                    n.set(add)
                    C=add
                elif(var3.get()==0):
                    n.set("")

            def fatname():
                if(var4.get()==1):
                    fatnam=fathers_name.get()
                    o.set(fatnam)
                    D=fatnam
                elif(var4.get()==0):
                    o.set("")

            def motname():
                if(var5.get()==1):
                    motnam=mothers_name.get()
                    p.set(motnam)
                    E=motnam
                elif(var5.get()==0):
                    p.set("")
            
            def gen1():
                if(var6.get()==1):
                    gen=gender.get()
                    q.set(gen)
                    F=gen
                elif(var6.get()==0):
                    q.set("")

            def birth1():
                if(var7.get()==1):
                    birth=birthdate.get()
                    r.set(birth)
                    G=birth
                elif(var7.get()==0):
                    r.set("")

            def maristat1():
                if(var8.get()==1):
                    maristat=marital_status.get()
                    s.set(maristat)
                    H=maristat
                elif(var8.get()==0):
                    s.set("")

            def rlg1():
                if(var9.get()==1):
                    rlg=religion.get()
                    t.set(rlg)
                    I=rlg
                elif(var9.get()==0):
                    t.set("")

            def lng1():
                if(var10.get()==1):
                    lng=language.get()
                    u.set(lng)
                    J=lng
                elif(var10.get()==0):
                    u.set("")

            def qlfct1():
                if(var11.get()==1):
                    qlfct=qualification.get()
                    v.set(qlfct)
                    K=qlfct
                elif(var11.get()==0):
                    v.set("")

            def exp1():
                if(var12.get()==1):
                    exp=experience.get()
                    w.set(exp)
                    L=exp
                elif(var12.get()==0):
                    w.set("")

            def addr1():
                if(var13.get()==1):
                    addr=address.get()
                    x.set(addr)
                    M=addr
                elif(var13.get()==0):
                    x.set("")

            def plc1():
                if(var14.get()==1):
                    plc=place.get()
                    y.set(plc)
                    N=plc
                elif(var14.get()==0):
                    y.set("")
            
            def dat1():
                if(var15.get()==1):
                    dat=date.get()
                    z.set(dat)
                    O=dat
                elif(var15.get()==0):
                    z.set("")
                
            label_name=tkinter.Label(top,text="New Name",font=("bold",14),pady=(10))
            label_name.grid(row=1,column=1)

            l = StringVar()
            m = StringVar()
            n = StringVar()
            o = StringVar()
            p = StringVar()
            q = StringVar()
            r = StringVar()
            s = StringVar()
            t = StringVar()
            u = StringVar()
            v = StringVar()
            w = StringVar()
            x = StringVar()
            y = StringVar()
            z = StringVar()

            
            
            name1=Entry(top, width=25, font=('Helvetica',14), textvariable=l)
            name1.grid(row=2,column=1,padx=10)

            var1 = IntVar()
            c1=Checkbutton(top, text="same as before",variable=var1,command=nam)
            c1.grid(row=3,column=1)
            
            label_mobile=tkinter.Label(top,text="New Mobile Number",font=("bold",14),pady=(10))
            label_mobile.grid(row=4,column=1)

            mobile1=Entry(top, width=25, font=('Helvetica',14), textvariable=m)
            mobile1.grid(row=5,column=1)

            var2 = IntVar()
            c2=Checkbutton(top, text="same as before", variable=var2,command=ph)
            c2.grid(row=6,column=1)
                
            label_email=tkinter.Label(top,text="New Email id",font=("bold",14),pady=(10))
            label_email.grid(row=7,column=1)

            email1=Entry(top, width=25, font=('Helvetica',14), textvariable=n)
            email1.grid(row=8,column=1)

            var3 = IntVar()
            c3=Checkbutton(top, text="same as before", variable=var3,command=ad)
            c3.grid(row=9,column=1)

            label_father=tkinter.Label(top,text="New Father's Name",font=("bold",14),pady=(10))
            label_father.grid(row=10,column=1)

            father1=Entry(top, width=25, font=('Helvetica',14), textvariable=o)
            father1.grid(row=11,column=1)

            var4 = IntVar()
            c4=Checkbutton(top, text="same as before", variable=var4,command=fatname)
            c4.grid(row=12,column=1)

            label_mother=tkinter.Label(top,text="New Mother's Name",font=("bold",14),pady=(10))
            label_mother.grid(row=13,column=1)

            mother1=Entry(top, width=25, font=('Helvetica',14), textvariable=p)
            mother1.grid(row=14,column=1)

            var5 = IntVar()
            c5=Checkbutton(top, text="same as before", variable=var5,command=motname)
            c5.grid(row=15,column=1)

            label_gender=tkinter.Label(top,text="New Gender",font=("bold",14),pady=(10))
            label_gender.grid(row=1,column=2)

            gender1=Entry(top, width=25, font=('Helvetica',14), textvariable=q)
            gender1.grid(row=2,column=2)

            var6 = IntVar()
            c6=Checkbutton(top, text="same as before", variable=var6,command=gen1)
            c6.grid(row=3,column=2)

            label_birthdate=tkinter.Label(top,text="New Date of Birth",font=("bold",14),pady=(10))
            label_birthdate.grid(row=4,column=2)

            birthdate1=Entry(top, width=25, font=('Helvetica',14), textvariable=r)
            birthdate1.grid(row=5,column=2,padx=10)

            var7 = IntVar()
            c7=Checkbutton(top, text="same as before", variable=var7,command=birth1)
            c7.grid(row=6,column=2)

            label_maritalstatus=tkinter.Label(top,text="New Marital Status",font=("bold",14),pady=(10))
            label_maritalstatus.grid(row=7,column=2)

            maritalstatus1=Entry(top, width=25, font=('Helvetica',14), textvariable=s)
            maritalstatus1.grid(row=8,column=2)

            var8 = IntVar()
            c8=Checkbutton(top, text="same as before", variable=var8,command=maristat1)
            c8.grid(row=9,column=2)

            label_religion=tkinter.Label(top,text="New Religion",font=("bold",14),pady=(10))
            label_religion.grid(row=10,column=2)

            religion1=Entry(top, width=25, font=('Helvetica',14), textvariable=t)
            religion1.grid(row=11,column=2)

            var9 = IntVar()
            c9=Checkbutton(top, text="same as before", variable=var9,command=rlg1)
            c9.grid(row=12,column=2)

            label_language=tkinter.Label(top,text="New Languages Known",font=("bold",14),pady=(10))
            label_language.grid(row=13,column=2)

            language1=Entry(top, width=25, font=('Helvetica',14), textvariable=u)
            language1.grid(row=14,column=2)

            var10 = IntVar()
            c10=Checkbutton(top, text="same as before", variable=var10,command=lng1)
            c10.grid(row=15,column=2)

            label_qualification=tkinter.Label(top,text="New Qualification",font=("bold",14),pady=(10))
            label_qualification.grid(row=1,column=3)

            qualification1=Entry(top, width=25, font=('Helvetica',14), textvariable=v)
            qualification1.grid(row=2,column=3)

            var11 = IntVar()
            c11=Checkbutton(top, text="same as before", variable=var11,command=qlfct1)
            c11.grid(row=3,column=3)

            label_experience=tkinter.Label(top,text="New Experience",font=("bold",14),pady=(10))
            label_experience.grid(row=4,column=3)

            experience1=Entry(top, width=25, font=('Helvetica',14), textvariable=w)
            experience1.grid(row=5,column=3,padx=10)

            var12 = IntVar()
            c12=Checkbutton(top, text="same as before", variable=var12,command=exp1)
            c12.grid(row=6,column=3)

            label_address=tkinter.Label(top,text="New Address",font=("bold",14),pady=(10))
            label_address.grid(row=7,column=3)

            address1=Entry(top, width=25, font=('Helvetica',14), textvariable=x)
            address1.grid(row=8,column=3)

            var13 = IntVar()
            c13=Checkbutton(top, text="same as before", variable=var13,command=addr1)
            c13.grid(row=9,column=3)

            label_place=tkinter.Label(top,text="New Place",font=("bold",14),pady=(10))
            label_place.grid(row=10,column=3)

            place1=Entry(top, width=25, font=('Helvetica',14), textvariable=y)
            place1.grid(row=11,column=3)

            var14 = IntVar()
            c14=Checkbutton(top, text="same as before", variable=var14,command=plc1)
            c14.grid(row=12,column=3)
            
            label_date=tkinter.Label(top,text="New Date",font=("bold",14),pady=(10))
            label_date.grid(row=13,column=3)

            date1=Entry(top, width=25, font=('Helvetica',14), textvariable=z)
            date1.grid(row=14,column=3)

            var15 = IntVar()
            c15=Checkbutton(top, text="same as before", variable=var15,command=dat1)
            c15.grid(row=15,column=3)

            def update():
                print(ws['A'+str(each_cell)])
                ws['A'+str(each_cell)] = name1.get()
                ws['B'+str(each_cell)] = mobile1.get()
                ws['C'+str(each_cell)] = email1.get()
                ws['D'+str(each_cell)] = father1.get()
                ws['E'+str(each_cell)] = mother1.get()
                ws['F'+str(each_cell)] = gender1.get()
                ws['G'+str(each_cell)] = birthdate1.get()
                ws['H'+str(each_cell)] = maritalstatus1.get()
                ws['I'+str(each_cell)] = religion1.get()
                ws['J'+str(each_cell)] = language1.get()
                ws['K'+str(each_cell)] = qualification1.get()
                ws['L'+str(each_cell)] = experience1.get()
                ws['M'+str(each_cell)] = address1.get()
                ws['N'+str(each_cell)] = place1.get()
                ws['O'+str(each_cell)] = date1.get()
                messagebox.showinfo("The information was updated successfully")
                clear_entries()
                wb.save('biodata.xlsx')


            Bupdate=tkinter.Button(top,text="Update",command=update, bg='red', font=("Helevetica",11,"bold"))
            Bupdate.grid(row=16,column=2)  
            
            top.mainloop()
        
        
                
    else:
        result_add.config(text="no record found to update!")

button3=Button(root,text="Search",width=10,height=2,bg='lightblue', font=("Helevetica",11,"bold"),command=lambda:search())
button3.place(x=240,y=620)

def search():
    for each_cell in range(2,(ws.max_row)+1):
        if (name.get() == ws['A'+str(each_cell)].value) or (mobile.get() == ws['B'+str(each_cell)].value):  
            Found =  True;
            cell_address = str(each_cell)
            break;
        else:
            Found=False
    if(Found == True):
        messagebox.showinfo("Found","Data Exist IN" + cell_address)
        mobile.insert(0,ws['B'+cell_address].value)
        email.insert(0,ws['C'+cell_address].value) 
        fathers_name.insert(0,ws['D'+cell_address].value)
        mothers_name.insert(0,ws['E'+cell_address].value) 
        gender.insert(0,ws['F'+cell_address].value)
        birthdate.insert(0,ws['G'+cell_address].value) 
        marital_status.insert(0,ws['H'+cell_address].value)
        religion.insert(0,ws['I'+cell_address].value) 
        language.insert(0,ws['J'+cell_address].value)
        qualification.insert(0,ws['K'+cell_address].value) 
        experience.insert(0,ws['L'+cell_address].value)
        address.insert(0,ws['M'+cell_address].value) 
        place.insert(0,ws['N'+cell_address].value) 
        date.insert(0,ws['O'+cell_address].value) 

button4=Button(root,text="Delete",width=10,height=2,bg='orange', font=("Helevetica",11,"bold"),command=lambda:delete())
button4.place(x=350,y=620)

def delete():
    for each_cell in range(2,(ws.max_row)+1):
        if (name.get() == ws['A'+str(each_cell)].value) or (mobile.get() == ws['B'+str(each_cell)].value):  
            Found =  True;
            cell_address = (each_cell)
            break;
        else:
            Found=False
    if(Found == True):
        ws.delete_rows(cell_address)
        messagebox.showinfo("Info","Data Deleted")
        clear_entries()
    wb.save('biodata.xlsx')

button5=Button(root,text="Show Result",width=10,height=2,bg='green', font=("Helevetica",11,"bold"),command=lambda:show_result())
button5.place(x=460,y=620)

def show_result():
    result_name.config(text="")
    result_mobile.config(text="")
    result_email.config(text="")
    result_fathers_name.config(text="")
    result_mothers_name.config(text="")
    result_gender.config(text="")
    result_birthdate.config(text="")
    result_marital_status.config(text="")
    result_religion.config(text="")
    result_language.config(text="")
    result_qualification.config(text="")
    result_experience.config(text="")
    result_address.config(text="")
    result_place.config(text="")
    result_date.config(text="")

    for each_cell in range(2,(ws.max_row)+1):
        if((name.get()==ws['A'+str(each_cell)].value) or (mobile.get()==ws['B'+str(each_cell)].value)):
            Found=True
            break

        else:
            Found = False  
            
    if (Found==True):
        result_name.config(text="Name: "+str(ws['A'+str(each_cell)].value))
        result_mobile.config(text="Mobile: "+str(ws['B'+str(each_cell)].value))
        result_email.config(text="Email id: "+str(ws['C'+str(each_cell)].value))
        result_fathers_name.config(text="Father's Name: "+str(ws['D'+str(each_cell)].value))
        result_mothers_name.config(text="Mother's Name: "+str(ws['E'+str(each_cell)].value))
        result_gender.config(text="Gender: "+str(ws['F'+str(each_cell)].value))
        result_birthdate.config(text="Date of Birth: "+str(ws['G'+str(each_cell)].value))
        result_marital_status.config(text="Marital Status: "+str(ws['H'+str(each_cell)].value))
        result_religion.config(text="Religion: "+str(ws['I'+str(each_cell)].value))
        result_language.config(text="Languages Known: "+str(ws['J'+str(each_cell)].value))
        result_qualification.config(text="Qualification: "+str(ws['K'+str(each_cell)].value))
        result_experience.config(text="Experience: "+str(ws['L'+str(each_cell)].value))
        result_address.config(text="Address: "+str(ws['M'+str(each_cell)].value))
        result_place.config(text="Place: "+str(ws['N'+str(each_cell)].value)) 
        result_date.config(text="Date: "+str(ws['O'+str(each_cell)].value))

    else:
        result_add.config(text="No record found.")  

column_A = ws['A']
column_B = ws['B']
column_C = ws['C']
column_D = ws['D']
column_E = ws['E']
column_F = ws['F']
column_G = ws['G']
column_H = ws['H']
column_I = ws['I']
column_J = ws['J']
column_K = ws['K']
column_L = ws['L']
column_M = ws['M']
column_N = ws['N']
column_O = ws['O']

def clear_entries():
    name.delete(0, END)
    mobile.delete(0, END)
    email.delete(0, END)
    fathers_name.delete(0, END)
    mothers_name.delete(0, END)
    gender.delete(0, END)
    birthdate.delete(0, END)
    marital_status.delete(0, END)
    religion.delete(0, END)
    language.delete(0, END)
    qualification.delete(0, END)
    experience.delete(0, END)
    address.delete(0, END)
    place.delete(0, END)
    date.delete(0, END)

label_result=Label(root,text="Result",font=("bold",17,"underline"))
label_result.grid(row=0,column=4,padx=250)

result_name=Label(root,font=("bold,14"))
result_name.grid(row=1,column=4,padx=150,sticky=W)

result_mobile=Label(root,font=("bold,14"))
result_mobile.grid(row=2,column=4,padx=150,sticky=W)

result_email=Label(root,font=("bold,14"))
result_email.grid(row=3,column=4,padx=150,sticky=W)

result_fathers_name=Label(root,font=("bold,14"))
result_fathers_name.grid(row=4,column=4,padx=150,sticky=W)

result_mothers_name=Label(root,font=("bold,14"))
result_mothers_name.grid(row=5,column=4,padx=150,sticky=W)

result_gender=Label(root,font=("bold,14"))
result_gender.grid(row=6,column=4,padx=150,sticky=W)

result_birthdate=Label(root,font=("bold,14"))
result_birthdate.grid(row=7,column=4,padx=150,sticky=W)

result_marital_status=Label(root,font=("bold,14"))
result_marital_status.grid(row=8,column=4,padx=150,sticky=W)

result_religion=Label(root,font=("bold,14"))
result_religion.grid(row=9,column=4,padx=150,sticky=W)

result_language=Label(root,font=("bold,14"))
result_language.grid(row=10,column=4,padx=150,sticky=W)

result_qualification=Label(root,font=("bold,14"))
result_qualification.grid(row=11,column=4,padx=150,sticky=W)

result_experience=Label(root,font=("bold,14"))
result_experience.grid(row=12,column=4,padx=150,sticky=W)

result_address=Label(root,font=("bold,14"))
result_address.grid(row=13,column=4,padx=150,sticky=W)

result_place=Label(root,font=("bold,14"))
result_place.grid(row=14,column=4,padx=150,sticky=W)

result_date=Label(root,font=("bold,14"))
result_date.grid(row=15,column=4,padx=150,sticky=W)

result_add=Label(root,font=("bold,14"))
result_add.grid(row=9,column=4,padx=150)

root.mainloop()
