from ast import Pass
from cgitb import text
from tkinter import *
from tkinter import font
from tkinter import messagebox 
from openpyxl.workbook import Workbook 
from openpyxl import load_workbook

from sample_registration import show_result 

root = Tk()
root.geometry('720x500')
#root.iconbitmap('d:/PYTHON PROJECTS/CC103/registration.ico')
root.title('sample registration page')
root.configure(background="lightblue")
wb = Workbook()
wb = load_workbook('register1.xlsx')
ws = wb['Sheet1']

RegFrmLbl = Label(root, text="REGISTRATION FORM", font=('Arial',20))

FirstNameLabel = Label(root, text="FIRST NAME",font=('Helvetica', 14) )
FirstName = Entry(root, width=40, font=('Helvetica', 14))

LastNameLabel = Label(root, text="LAST NAME",font=('Helvetica', 14) )
LastName = Entry(root, width=40, font=('Helvetica',14))

EmailAddressLabel = Label(root, text="EMAIL ADDRESS",font=('Helvetica',14))
EmailAddress = Entry(root, width=40, font=('Helvetica',14))

SchoolDepartmentLabel = Label(root, text="Department", font=('Helvetica',14))
selected_department = IntVar()
Department1 = Radiobutton(root, text='BSIT', variable=selected_department,value=1,font=('Helvetica',17))
Department2 = Radiobutton(root, text='BSPA', variable=selected_department,value=2,font=('Helvetica',17))
Department3 = Radiobutton(root, text='BSEntrep', variable=selected_department,value=3,font=('Helvetica',17))
Department4 = Radiobutton(root, text='BSSW', variable=selected_department, value=4,font=('Helvetica',17))

CityLabel = Label(root, text="CITY / TOWN", font=('Helvetica',14))
selected_city = StringVar(root)
city_list = ["Lucena City", "Sariaya", "Pagbilao", "Tayabas City", "Candelaria", "Padre Burgos","Lucban","Gumaca","Unisan","Other Parts of Quezon"]
selected_city.set(city_list[0])

City = OptionMenu(root, selected_city,*city_list)

PasswordLabel = Label(root, text="PASSWORD", font=('Helvetica',14))
Password = Entry(root, show="*", width=40, font=('Helvetica',14))

RegisterBtn = Button(root, text="Register",width=15,height=4,font=('Arial',15),command=lambda:save_data())
ViewRecordsBtn = Button(root, text="View Records",width=15,height=4,font=('Arial',15), command=lambda:show_result())
dsearch=Button(root,text="Search", width=15, height=4, font=("Arial",15), command=lambda:search_data())

def save_data():
    #check whether the data is already in the database/excel
    for each_cell in range(2, (ws.max_row) + 1):
        if (FirstNameLabel.get() == ws['A' + str(each_cell)].value) or (LastNameLabel.get() == ws['B' + str(each_cell)].value):
            Found = True;
            break;
        else:
            Found = False
    if(Found == True):
        messagebox.showinfo("ERROR","DATA ALREADY EXISTED")
    else:
        lastRow = str(ws.max_row + 1)
        ws['A' + lastRow] = FirstNameLabel.get() 
        ws['B' + lastRow] = LastNameLabel.get()
        ws['C' + lastRow] = EmailAddressLabel.get()
        ws['D' + lastRow] = SchoolDepartmentLabel.get()
        ws['E' + lastRow] = CityLabel.get()
        ws['F' + lastRow] = PasswordLabel.get()
   
        messagebox.showinfo("SUCCESS","DATA SAVED SUCCESSFULLY")
    wb.save('register2.xlsx')

def search_data():
    for each_cell in range(2, (ws.max_row) + 1):
        if (FirstNameLabel.get() == ws['A' + str(each_cell)].value) or (LastNameLabel.get() == ws['B' + str(each_cell)].value):
            Found = True;
            cell_address = str(each_cell)
        else:
            Found = False
    if(Found == True):
        messagebox.info("FOUND","DATA EXIST IN " + cell_address)
        LastNameLabel.insert(0,ws['B' + cell_address].value)
        EmailAddressLabel.insert(0,ws['C' + cell_address].value)
        SchoolDepartmentLabel.insert(0,ws['D' + cell_address].value)
        CityLabel.insert(0,ws['E' + cell_address].value)
        PasswordLabel.insert(0,ws['F' + cell_address].value)
column_A = ws['A']
column_B = ws['B']
column_C = ws['C']
column_D = ws['D']
column_E = ws['E']
column_F = ws['F']

results_show = Label(root,font=('Helvetica',14))

def show_result():
    results_list = ""
    for each_cell in column_A:
        #print(each_cell.value)
        results_list = results_list + each_cell.value    
        results_list = f'{results_list + str(each_cell.value)}\n'
    results_show.configure(text=results_list)

#Grid Layout
RegFrmLbl.grid(row=0,column=0, columnspan=2,padx=3,pady=1)

#row1
FirstNameLabel.grid(row=1,column=0,padx=3,pady=3)
FirstName.grid(row=1, column=1, padx=3,pady=3)

#row2
LastNameLabel.grid(row=2,column=0,padx=3,pady=3)
LastName.grid(row=2,column=1,padx=3,pady=3)

#row3
EmailAddressLabel.grid(row=3,column=0,padx=3,pady=3)
EmailAddress.grid(row=3,column=1,padx=3,pady=3)

#row4 - 
SchoolDepartmentLabel.grid(row=4,column=0,padx=3,pady=3)
Department1.grid(row=4,column=1, padx=3,pady=3,sticky=W)

#row5
Department2.grid(row=5,column=1,padx=3,pady=3,sticky=W)

#row6
Department3.grid(row=6,column=1,padx=3,pady=3,sticky=W)

#row7
Department4.grid(row=7,column=1,padx=3,pady=3,sticky=W)

#row8
CityLabel.grid(row=8,column=0,padx=3,pady=3)
City.grid(row=8,column=1,padx=3,pady=3,sticky=W)


#row9
PasswordLabel.grid(row=9,column=0,padx=3,pady=3)
Password.grid(row=9,column=1,padx=3,pady=3)

#row10
RegisterBtn.grid(row=10,column=0,padx=3,pady=3)
ViewRecordsBtn.grid(row=10, column=1,padx=3,pady=3)
dsearch.grid(row=10, column=2,padx=3,pady=3)
root.mainloop()