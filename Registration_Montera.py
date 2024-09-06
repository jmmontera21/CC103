from tkinter import *
import os
from tkinter import messagebox
from openpyxl.workbook import Workbook
from openpyxl import load_workbook

#NOTE: ONLY IF NEEDED 
# os.chdir('C:\\Users\\leona\\OneDrive\\Desktop\\CC103\\CRUD')
# os.getcwd()  

window=Tk()
window.title("Personal Information")
window.geometry("500x900")


wb = Workbook()
wb=load_workbook('register.xlsx')
ws = wb.active

label=Label(window,text="Register",font=("bold",20), fg='blue', pady=(20))
label.pack()

label_name=Label(window,text="Name",font=("bold",14),pady=(20))
label_name.pack()

name=Entry(window, width=30, font=('Helvetica',14))
name.pack()

  
label_phone=Label(window,text="Phone",font=("bold",14),pady=(20))
label_phone.pack()

phone=Entry(window,width=30, font=('Helvetica',14))
phone.pack()

label_add=Label(window,text="Address",font=("bold",14),pady=(20))
label_add.pack()
add1=Entry(window,width=30, font=('Helvetica',14))
add1.pack()

addbtn=Button(window,text="Add",width=10, height=1, bg='Yellow',font=("Helevetica",11,"bold"), pady=2, command=lambda:save_data())

def save_data():
    for each_cell in range(2,(ws.max_row)+1):
        if (name.get() == ws['A'+str(each_cell)].value) or (phone.get() == ws['B'+str(each_cell)].value):  
            Found =  True;
            break;
        else:
            Found = False
    if(Found ==True):
        messagebox.showinfo("ERROR","Data Already Exist")
    else:
        lastrow = str((ws.max_row)+1)
        ws['A'+lastrow] = name.get()
        ws['B'+lastrow] = phone.get()
        ws['C'+lastrow] = add1.get()   

        messagebox.showinfo("SUCCESS","Data Saved Successfully")
    wb.save('register.xlsx')

               
search=Button(window,text="Search",width=10, height=1, bg='pink', font=("Helevetica",11,"bold"),command=lambda:search_data())

def search_data():
    for each_cell in range(2,(ws.max_row)+1):
        if (name.get() == ws['A'+str(each_cell)].value) or (phone.get() == ws['B'+str(each_cell)].value):  
            Found =  True;
            cell_address = str(each_cell)
            break;
        else:
            Found=False
    if(Found == True):
        messagebox.showinfo("Found","Data Exist IN" + cell_address)
        phone.insert(0,ws['B'+cell_address].value)
        add1.insert(0,ws['C'+cell_address].value) 

def clear_entry():
    name.delete(0,END)
    phone.delete(0,END)
    add1.delete(0,END)

showresult=Button(window,text="Show Result",width=10, height=1, bg='orange', font=("Helevetica",11,"bold"),command=lambda:show_result())


column_A = ws['A']
column_B = ws['B']
column_C = ws['C']

results_show = Label(window, font=("Helevetica",14))

def show_result():
    result_name.config(text="")
    result_phone.config(text="")
    result_add1.config(text="")

    for each_cell in range(2,(ws.max_row)+1):
        if (name.get() == ws['A'+str(each_cell)].value) or (phone.get() == ws['B'+str(each_cell)].value):  
            Found =  True;
            cell_address = str(each_cell)
            break;
        else:
            Found=False
    if(Found == True):
        messagebox.showinfo("Found","Data Exist IN" + cell_address)
        result_name.config(text="Name: "+str(ws['A'+str(each_cell)].value))
        result_phone.config(text="Phone: "+str(ws['B'+str(each_cell)].value))
        result_add1.config(text="Address: "+str(ws['C'+str(each_cell)].value)) 
           
addbtn.pack()
search.pack()
showresult.pack()

result_name=Label(window,font=("bold",15))
result_name.pack()

result_phone=Label(window,font=("bold",15))
result_phone.pack()

result_add1=Label(window,font=("bold",15))
result_add1.pack()

res=Label(window,text="Result",font=("Arial", 17, "bold"),pady=20)
res.pack()



window.mainloop()